# -*- coding: utf-8 -*-
import base64
import io
import zipfile
from datetime import datetime, timezone
from odoo import api, fields, models, _
from odoo.exceptions import UserError


class LicenseProviderPartner(models.Model):
    _name = 'license.provider.partner'
    _description = 'Proveedor de licencias (contacto seleccionado)'
    _order = 'partner_id'

    partner_id = fields.Many2one(
        'res.partner',
        string='Proveedor',
        required=True,
        ondelete='cascade',
        help='Contacto que actúa como proveedor de licencias.',
    )
    partner_name = fields.Char(related='partner_id.name', string='Nombre', readonly=True)
    partner_phone = fields.Char(related='partner_id.phone', string='Teléfono', readonly=True)
    partner_city = fields.Char(related='partner_id.city', string='Ciudad', readonly=True)
    cut_off_date = fields.Date(
        string='Fecha de Corte',
        help='Fecha de corte de referencia del proveedor. La fecha de corte que aplica por cliente es la misma que las fechas y condiciones que se asignan al cliente en cada asignación (fecha inicio/fin y condiciones del contrato).',
    )
    provided_licenses_count = fields.Integer(
        string='Licencias Proporcionadas',
        compute='_compute_provided_licenses_count',
        help='Cantidad de tipos de licencia (productos) que este proveedor entrega. Ver detalle con el botón.',
    )
    assignments_cut_off_count = fields.Integer(
        string='Asignaciones',
        compute='_compute_assignments_cut_off_count',
        help='Asignaciones a clientes que usan este proveedor.',
    )
    assignment_ids = fields.Many2many(
        'license.assignment',
        string='Licencias proporcionadas',
        compute='_compute_assignment_ids',
        help='Asignaciones a clientes con este proveedor (tabla tipo Excel).',
    )
    report_line_ids = fields.One2many(
        'license.provider.report.line',
        'provider_partner_id',
        string='Líneas de reporte / facturación',
        help='Datos cargados desde los reportes del proveedor (Excel). Cada proveedor puede tener formato distinto; aquí se unifica la información principal.',
    )
    report_group_ids = fields.One2many(
        'license.provider.report.group',
        'provider_partner_id',
        string='Clientes en reporte',
        help='Un registro por cliente; use «Ver licencias contratadas» para ver el detalle.',
    )
    stock_ids = fields.One2many(
        'license.provider.stock',
        'provider_partner_id',
        string='Licencias',
        help='Licencias y costos por licencia de este proveedor. Al agregar una línea desde aquí se guarda correctamente.',
    )

    @api.depends('partner_id')
    def _compute_assignment_ids(self):
        Assignment = self.env['license.assignment']
        for rec in self:
            if rec.partner_id:
                rec.assignment_ids = Assignment.search([
                    ('license_provider_id', '=', rec.partner_id.id),
                    ('state', 'in', ['draft', 'active']),
                ])
            else:
                rec.assignment_ids = Assignment.browse([])

    @api.depends('partner_id')
    def _compute_provided_licenses_count(self):
        Stock = self.env['license.provider.stock']
        for rec in self:
            if rec.partner_id:
                rec.provided_licenses_count = Stock.search_count([('provider_id', '=', rec.partner_id.id)])
            else:
                rec.provided_licenses_count = 0

    @api.depends('partner_id')
    def _compute_assignments_cut_off_count(self):
        Assignment = self.env['license.assignment']
        for rec in self:
            if rec.partner_id:
                rec.assignments_cut_off_count = Assignment.search_count([
                    ('license_provider_id', '=', rec.partner_id.id),
                    ('state', 'in', ['draft', 'active']),
                ])
            else:
                rec.assignments_cut_off_count = 0

    def action_view_assignments_cut_off(self):
        """Abre las asignaciones de este proveedor. La fecha de corte por cliente es la misma que las fechas/condiciones de cada asignación."""
        self.ensure_one()
        return {
            'name': _('Asignaciones (fecha de corte = condiciones del cliente) - %s') % (self.partner_id.name or 'Proveedor'),
            'type': 'ir.actions.act_window',
            'res_model': 'license.assignment',
            'view_mode': 'list,form',
            'domain': [
                ('license_provider_id', '=', self.partner_id.id),
                ('state', 'in', ['draft', 'active']),
            ],
            'context': {'default_license_provider_id': self.partner_id.id},
        }

    def action_sync_stock_from_provider(self):
        """Enlaza a este proveedor las licencias que tienen provider_id = este contacto pero aún no tienen provider_partner_id."""
        self.ensure_one()
        if not self.partner_id:
            return
        Stock = self.env['license.provider.stock']
        orphans = Stock.search([
            ('provider_id', '=', self.partner_id.id),
            ('provider_partner_id', '=', False),
        ])
        if orphans:
            orphans.write({'provider_partner_id': self.id})

    def action_view_provided_licenses(self):
        """Abre el stock de licencias de este proveedor (varias licencias distintas)."""
        self.ensure_one()
        return {
            'name': _('Licencias de %s') % (self.partner_id.name or 'Proveedor'),
            'type': 'ir.actions.act_window',
            'res_model': 'license.provider.stock',
            'view_mode': 'list,form',
            'domain': [('provider_id', '=', self.partner_id.id)],
            'context': {'default_provider_id': self.partner_id.id},
        }

    def _notify(self, message, notification_type='success'):
        """Devuelve una acción cliente para mostrar una notificación toast en la interfaz."""
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Aviso'),
                'message': message,
                'type': notification_type,
                'sticky': False,
            },
        }

    def action_sync_report_groups(self):
        """Sincroniza líneas de reporte desde asignaciones (crea/actualiza sin tocar costo) y luego actualiza la lista de clientes.
        Si no hay asignaciones, elimina líneas huérfanas (asignación pasada a otro proveedor) y actualiza la lista."""
        self.ensure_one()
        ReportLine = self.env['license.provider.report.line']
        Assignment = self.env['license.assignment']
        assignments = Assignment.search([
            ('license_provider_id', '=', self.partner_id.id),
            ('state', 'in', ['draft', 'active']),
        ])
        if not assignments:
            lines_with_assignment = ReportLine.search([
                ('provider_partner_id', '=', self.id),
                ('assignment_id', '!=', False),
            ])
            orphans = lines_with_assignment.filtered(
                lambda l: l.assignment_id.license_provider_id != self.partner_id
            )
            if orphans:
                orphans.unlink()
            self.invalidate_recordset(['report_line_ids', 'report_group_ids'])
            self._sync_report_groups()
            return self._notify(_('No hay asignaciones con este proveedor. Se quitaron los clientes que ya no lo tienen.'), notification_type='warning')
        for assig in assignments:
            self._sync_report_line_for_assignment(assig, sync_groups=False)
        self.invalidate_recordset(['report_line_ids', 'report_group_ids'])
        self._sync_report_groups()
        return self._notify(_('Lista actualizada: %s asignación(es) sincronizada(s).') % len(assignments))

    def _sync_report_groups(self):
        """Crea o actualiza un registro por cliente en report_group_ids a partir de report_line_ids.
        Solo se cuentan líneas cuya asignación (si tiene) sigue teniendo este proveedor."""
        self.ensure_one()
        Group = self.env['license.provider.report.group']
        partners_with_lines = set()
        for line in self.report_line_ids:
            if line.assignment_id and line.assignment_id.license_provider_id != self.partner_id:
                continue
            if not line.partner_id:
                continue
            if line.partner_id.id in partners_with_lines:
                continue
            partners_with_lines.add(line.partner_id.id)
            name = line.client_name or line.partner_id.name or _('Sin nombre')
            existing = Group.search([
                ('provider_partner_id', '=', self.id),
                ('partner_id', '=', line.partner_id.id),
            ], limit=1)
            if existing:
                existing.client_name = name
            else:
                Group.create({
                    'provider_partner_id': self.id,
                    'partner_id': line.partner_id.id,
                    'client_name': name,
                })
        # Eliminar grupos cuyo cliente ya no tiene líneas
        for group in self.report_group_ids:
            if group.partner_id.id not in partners_with_lines:
                group.unlink()

    def _map_contracting(self, ct):
        if not ct:
            return 'monthly_monthly', 'Mensual'
        if ct == 'annual':
            return 'annual', 'Anual'
        if ct == 'annual_monthly_commitment':
            return 'annual_monthly_commitment', 'Anual Compromiso Mensual'
        return 'monthly_monthly', 'Mensual'

    def _report_line_vals_for_assignment(self, assig):
        """Construye vals para crear/actualizar una línea de reporte a partir de una asignación."""
        contract_type, billing_cycle = self._map_contracting(assig.contracting_type)
        unit_usd = (assig.total_cost_usd / assig.quantity) if assig.quantity else (assig.total_cost_usd or 0.0)
        return {
            'provider_partner_id': self.id,
            'partner_id': assig.partner_id.id if assig.partner_id else False,
            'client_name': assig.partner_id.name if assig.partner_id else '',
            'product_id': assig.license_id.product_id.id if assig.license_id and assig.license_id.product_id else False,
            'product_name': assig.license_id.product_id.name if assig.license_id and assig.license_id.product_id else (assig.license_display_name or ''),
            'quantity': assig.quantity,
            'start_date': assig.start_date,
            'end_date': assig.end_date,
            'cut_off_date': assig.end_date or assig.start_date,
            'contract_type': contract_type,
            'billing_cycle': billing_cycle,
            'unit_price_usd': unit_usd,
            'total_price_usd': assig.total_cost_usd or 0.0,
            'assignment_id': assig.id,
            'auto_renewal': assig.auto_renewal,
        }

    def _sync_report_line_for_assignment(self, assig, sync_groups=True):
        """Crea o actualiza una sola línea de reporte para la asignación. Al actualizar nunca escribe provider_cost_usd."""
        self.ensure_one()
        ReportLine = self.env['license.provider.report.line']
        existing = ReportLine.search([
            ('provider_partner_id', '=', self.id),
            ('assignment_id', '=', assig.id),
        ], limit=1)
        vals = self._report_line_vals_for_assignment(assig)
        if existing:
            update_vals = {k: v for k, v in vals.items() if k != 'provider_cost_usd'}
            existing.write(update_vals)
        else:
            vals['provider_cost_usd'] = 0.0
            ReportLine.create(vals)
        if sync_groups:
            self._sync_report_groups()

    def _excel_safe_str(self, val):
        """Convierte a string seguro para Excel (sin caracteres de control que corrompen el XML)."""
        if val is None:
            return ''
        s = str(val).strip()
        return ''.join(c for c in s if ord(c) >= 32 or c in '\t\n\r')

    def _make_core_xml_bytes(self):
        """Genera un docProps/core.xml mínimo y válido (OOXML) para que Excel no repare el archivo."""
        now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ').encode('utf-8')
        return (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
            b'xmlns:dc="http://purl.org/dc/elements/1.1/" '
            b'xmlns:dcterms="http://purl.org/dc/terms/" '
            b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            b'<dc:creator>Odoo</dc:creator>'
            b'<cp:lastModifiedBy>Odoo</cp:lastModifiedBy>'
            b'<dcterms:created xsi:type="dcterms:W3CDTF">' + now + b'</dcterms:created>'
            b'<dcterms:modified xsi:type="dcterms:W3CDTF">' + now + b'</dcterms:modified>'
            b'</cp:coreProperties>'
        )

    def _fix_xlsx_core_xml_for_excel(self, xlsx_bytes):
        """
        Sustituye docProps/core.xml por uno válido y quita xml:space="preserve" de todos los XML.
        Así Excel no detecta errores ni pide reparar el archivo.
        """
        try:
            core_xml = self._make_core_xml_bytes()
            with zipfile.ZipFile(io.BytesIO(xlsx_bytes), 'r') as zin:
                out_buf = io.BytesIO()
                with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for name in zin.namelist():
                        if name == 'docProps/core.xml':
                            data = core_xml
                        else:
                            data = zin.read(name)
                            if name.endswith('.xml'):
                                data = data.replace(b' xml:space="preserve"', b'')
                        zout.writestr(name, data)
                return out_buf.getvalue()
        except Exception:
            return xlsx_bytes

    def action_export_consolidated_excel(self):
        """Exporta las líneas del Consolidado (report_line_ids) a un archivo Excel."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, Border, Side
        except ImportError:
            raise UserError(_('Instale el paquete Python openpyxl: pip install openpyxl'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Consolidado'[:31]
        headers = [
            _('Cliente'), _('Producto / Oferta'), _('Fecha Inicio'), _('Fecha Fin'), _('Fecha Corte'),
            _('Contrato'), _('Cantidad'), _('Costo Proveedor'), _('Costo Total Proveedor'),
            _('Precio al Cliente'), _('Total Precio Cliente'), _('Ganancia'), _('Ganancia Total'),
            _('Renov. automática'),
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=self._excel_safe_str(h))
        header_font = Font(bold=True)
        thin = Side(style='thin')
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.border = Border(bottom=thin)
        # Evitar que openpyxl escriba xml:space="preserve" en core.xml (Excel lo rechaza)
        if hasattr(wb, 'properties') and wb.properties:
            for prop in ('creator', 'lastModifiedBy'):
                val = getattr(wb.properties, prop, None)
                if val is not None and isinstance(val, str):
                    setattr(wb.properties, prop, val.strip() or 'Odoo')
        contract_labels = dict(
            self.env['license.provider.report.line']._fields['contract_type'].selection or []
        )
        for row_idx, line in enumerate(self.report_line_ids, 2):
            client = line.client_name or (line.partner_id.name if line.partner_id else '') or ''
            contract = contract_labels.get(line.contract_type, '') or (line.contract_type or '')
            ws.cell(row=row_idx, column=1, value=self._excel_safe_str(client))
            ws.cell(row=row_idx, column=2, value=self._excel_safe_str(line.product_name))
            ws.cell(row=row_idx, column=3, value=line.start_date.strftime('%Y-%m-%d') if line.start_date else '')
            ws.cell(row=row_idx, column=4, value=line.end_date.strftime('%Y-%m-%d') if line.end_date else '')
            ws.cell(row=row_idx, column=5, value=line.cut_off_date.strftime('%Y-%m-%d') if line.cut_off_date else '')
            ws.cell(row=row_idx, column=6, value=self._excel_safe_str(contract))
            ws.cell(row=row_idx, column=7, value=int(line.quantity or 0))
            ws.cell(row=row_idx, column=8, value=float(line.provider_cost_usd or 0))
            ws.cell(row=row_idx, column=9, value=float(line.total_provider_cost_usd or 0))
            ws.cell(row=row_idx, column=10, value=float(line.unit_price_pricelist_usd or 0))
            ws.cell(row=row_idx, column=11, value=float(line.total_price_pricelist_usd or 0))
            ws.cell(row=row_idx, column=12, value=float(line.profit_unit_usd or 0))
            ws.cell(row=row_idx, column=13, value=float(line.profit_total_usd or 0))
            ws.cell(row=row_idx, column=14, value=self._excel_safe_str(_('Sí') if line.auto_renewal else _('No')))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        # Corregir docProps/core.xml: Excel espera dcterms:created/modified; openpyxl puede escribir dct:
        xlsx_bytes = self._fix_xlsx_core_xml_for_excel(buf.getvalue())
        name = _('Consolidado_%s.xlsx') % (self.partner_id.name or 'proveedor').replace('/', '-').replace('\\', '-')
        attach = self.env['ir.attachment'].create({
            'name': name,
            'type': 'binary',
            'datas': base64.b64encode(xlsx_bytes).decode('ascii'),
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=1' % attach.id,
            'target': 'self',
        }

    partner_uniq = models.Constraint(
        'unique(partner_id)',
        'Este contacto ya está en la lista de proveedores.',
    )
